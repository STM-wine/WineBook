import unittest
from pathlib import Path
from tempfile import TemporaryDirectory

from openpyxl import load_workbook

from grw_converter_app import (
    FileResolution,
    build_export_rows,
    build_optional_saasant_csv,
)
from modules.po_tools.grw_invoice_converter.grw_converter import write_to_updated_template
from modules.po_tools.grw_invoice_converter.parser import (
    extract_description_fragment_from_line,
    extract_invoice_summary_from_text,
    format_item_description,
    parse_item_block,
)
from modules.po_tools.grw_invoice_converter.validator import validate_no_duplicate_skus


class GrwParserTests(unittest.TestCase):
    def test_wrapped_descriptions_remain_distinct_for_duplicate_validation(self):
        block_one = (
            "5 Sale BUR:VIN:DANC- Vincent Dancer Chassagne Montrachet $149.00 1 750mL $149.00\n"
            "Premier Cru Tete du Clos Blanc 2017 750mL"
        )
        block_two = (
            "6 Sale BUR:VIN:DANC- Vincent Dancer Chassagne Montrachet $159.00 1 750mL $159.00\n"
            "La Romanee 2017 750mL"
        )

        item_one = parse_item_block(block_one)
        item_two = parse_item_block(block_two)

        self.assertIsNotNone(item_one)
        self.assertIsNotNone(item_two)
        self.assertIn("premier cru tete du clos blanc", item_one["clean_description"].lower())
        self.assertIn("la romanee", item_two["clean_description"].lower())
        self.assertNotEqual(item_one["clean_description"], item_two["clean_description"])

        validate_no_duplicate_skus([item_one, item_two])

    def test_wrapped_description_with_mixed_price_and_qty_text_is_preserved(self):
        block = (
            "5 Sale BUR:VIN:DANC- Vincent Dancer Chassagne Montrachet $149.00 1 750mL $149.00\n"
            "Premier Cru Tete du Clos Blanc 2017 750mL $149.00 1 750mL"
        )

        item = parse_item_block(block)

        self.assertIsNotNone(item)
        self.assertIn("premier cru tete du clos blanc", item["clean_description"].lower())
        self.assertIn("Vincent Dancer Chassagne Montrachet", item["description"])

    def test_code_prefixed_continuation_line_keeps_descriptive_remainder(self):
        fragment = extract_description_fragment_from_line(
            "0750-2017-F0L0C0 Premier Cru Tete du Clos Blanc 2017"
        )
        self.assertEqual(fragment, "Premier Cru Tete du Clos Blanc 2017")

        fragment_two = extract_description_fragment_from_line(
            "0750-2017-F0L0C0 La Romanee 2017 750mL"
        )
        self.assertEqual(fragment_two, "La Romanee 2017 750mL")

        code_only = extract_description_fragment_from_line("0750-1996-F0L0C0")
        self.assertEqual(code_only, "")

    def test_code_only_continuation_line_does_not_pollute_description(self):
        block = (
            "9 Sale BDX:CAN:GAFF- Canon La Gaffeliere OWC $300.00 1 PK12 $3,600.00\n"
            "PK12-2005-F0L0C0"
        )

        item = parse_item_block(block)

        self.assertIsNotNone(item)
        self.assertEqual(item["clean_description"], "Canon La Gaffeliere OWC")
        self.assertNotIn("F0L0C", item["description"])
        self.assertEqual(item["description"], "Canon La Gaffeliere OWC 2005 12/750ml")

    def test_code_fragments_are_removed_from_final_description(self):
        canon_block = (
            "9 Sale BDX:CAN:GAFF- Canon La Gaffeliere OWC -F0L0C0 $300.00 1 PK12 $3,600.00\n"
            "PK12-2005-F0L0C0"
        )
        faiveley_block = (
            "10 Sale BUR:FAI:MAZI- Faiveley Mazis Chambertin -F0L0C0 $250.00 1 750mL $250.00\n"
            "0750-2005-F0L0C0"
        )

        canon_item = parse_item_block(canon_block)
        faiveley_item = parse_item_block(faiveley_block)

        self.assertIsNotNone(canon_item)
        self.assertIsNotNone(faiveley_item)
        self.assertEqual(canon_item["description"], "Canon La Gaffeliere OWC 2005 12/750ml")
        self.assertEqual(faiveley_item["description"], "Faiveley Mazis Chambertin 2005 1/750ml")
        self.assertNotIn("F0L0C", canon_item["description"])
        self.assertNotIn("F0L0C", faiveley_item["description"])
        self.assertNotIn("--", canon_item["description"])
        self.assertNotIn("--", faiveley_item["description"])

    def test_one_point_five_liter_size_formats_cleanly(self):
        description = format_item_description(
            wine_name="Pol Roger Brut",
            vintage=2015,
            pack_size=3,
            sku_prefix="CHP",
            bottle_size="1.5L",
        )

        self.assertEqual(description, "Pol Roger Brut 2015 3/1500ml")

    def test_excel_export_defaults_item_number_column_to_new(self):
        items = [
            {
                "Item Number": "NEW",
                "Item Description": "Test Wine 2020 1/750ml",
                "Description": "Test Wine 2020 1/750ml",
                "PK": 1,
                "Quantity": 1,
                "FOB Btl": 10.0,
                "frontline": 12,
                "FOB Case": 10.0,
                "Ext Cost": 10.0,
                "Ext Price": 12.0,
                "SKU": "BUR",
                "STM Markup %": 0.10,
            }
        ]

        with TemporaryDirectory() as temp_dir:
            output_path = Path(temp_dir) / "test.xlsx"
            workbook_path = write_to_updated_template(
                items,
                "modules/po_tools/grw_invoice_converter/templates/GRW_Template_Updated.xlsx",
                str(output_path),
                "S12345",
                "Brix",
            )
            workbook = load_workbook(workbook_path)
            sheet = workbook.active

            self.assertEqual(sheet["A1"].value, "Item Number")
            self.assertEqual(sheet["A2"].value, "NEW")

    def test_excel_export_uses_finalized_item_description(self):
        items = [
            {
                "Item Number": "NEW",
                "Item Description": "Canon La Gaffeliere OWC 2005 12/750ml",
                "description": "RAW DESCRIPTION SHOULD NOT WIN",
                "PK": 12,
                "Quantity": 12,
                "FOB Btl": 25.0,
                "Frontline": 30,
                "FOB Case": 300.0,
                "Ext Cost": 300.0,
                "Ext Price": 360.0,
                "SKU": "BDX",
                "STM Markup %": 0.15,
            }
        ]

        with TemporaryDirectory() as temp_dir:
            output_path = Path(temp_dir) / "test.xlsx"
            workbook_path = write_to_updated_template(
                items,
                "modules/po_tools/grw_invoice_converter/templates/GRW_Template_Updated.xlsx",
                str(output_path),
                "S60490",
                "Brix",
            )
            workbook = load_workbook(workbook_path)
            sheet = workbook.active

            self.assertEqual(sheet["A2"].value, "NEW")
            self.assertEqual(sheet["B2"].value, "Canon La Gaffeliere OWC 2005 12/750ml")

    def test_export_rows_and_csv_share_same_clean_description(self):
        priced_items = [
            {
                "description": "Canon La Gaffeliere OWC 2024 12/750ml",
                "sku_prefix": "BDX",
                "pack_size": 12,
                "quantity": 12,
                "fob_bottle": 25.0,
                "frontline": 30,
                "fob_case": 300.0,
                "ext_cost": 300.0,
                "ext_price": 360.0,
            }
        ]
        resolution = FileResolution(customer_name="SNGC", invoice_number="S58725", used_fallback=False)

        export_rows = build_export_rows(priced_items, resolution)
        csv_filename, csv_bytes = build_optional_saasant_csv(export_rows, resolution, "SNGC_S58725_SAASANT.csv")

        self.assertEqual(export_rows[0]["Item Number"], "NEW")
        self.assertEqual(export_rows[0]["Item Description"], "Canon La Gaffeliere OWC 2024 12/750ml")
        self.assertNotIn("F0L0C", export_rows[0]["Item Description"])
        self.assertIn("Canon La Gaffeliere OWC 2024 12/750ml", csv_bytes.decode("utf-8"))
        self.assertIn("NEW", csv_bytes.decode("utf-8"))
        self.assertEqual(csv_filename, "SNGC_S58725_SAASANT.csv")

    def test_extract_invoice_summary_from_text_captures_credit_and_balance(self):
        text = """
        Date Payment Amount
        02/27/2026 Credit $ 1,553.75
        Subtotal: $1,700.00
        Sales Tax: $0.00
        Total: $1,700.00
        Paid: $1,553.75
        Balance Due: $146.25
        """

        summary = extract_invoice_summary_from_text(text)

        self.assertEqual(summary["credit_date"], "02/27/2026")
        self.assertEqual(summary["credit_amount"], 1553.75)
        self.assertEqual(summary["subtotal"], 1700.00)
        self.assertEqual(summary["paid_amount"], 1553.75)
        self.assertEqual(summary["balance_due"], 146.25)

    def test_excel_export_includes_invoice_adjustments_block(self):
        items = [
            {
                "Item Number": "NEW",
                "Item Description": "Leoville Poyferre 2003 1/750ml",
                "PK": 1,
                "Quantity": 5,
                "FOB Btl": 165.0,
                "Frontline": 190,
                "FOB Case": 165.0,
                "Ext Cost": 825.0,
                "Ext Price": 950.0,
                "SKU": "BDX",
                "STM Markup %": 0.15,
            }
        ]
        invoice_summary = {
            "subtotal": 1700.0,
            "credit_date": "02/27/2026",
            "credit_amount": 1553.75,
            "paid_amount": 1553.75,
            "balance_due": 146.25,
        }

        with TemporaryDirectory() as temp_dir:
            output_path = Path(temp_dir) / "test.xlsx"
            workbook_path = write_to_updated_template(
                items,
                "modules/po_tools/grw_invoice_converter/templates/GRW_Template_Updated.xlsx",
                str(output_path),
                "S59041",
                "Dan King",
                invoice_summary=invoice_summary,
            )
            workbook = load_workbook(workbook_path)
            sheet = workbook.active

            values = {
                (row, col): sheet.cell(row=row, column=col).value
                for row in range(1, 20)
                for col in range(1, 3)
            }

            self.assertIn("Invoice adjustments", values.values())
            self.assertIn("Credit Applied", values.values())
            self.assertIn("$1,553.75", values.values())
            self.assertIn("Balance Due", values.values())
            self.assertIn("$146.25", values.values())


if __name__ == "__main__":
    unittest.main()
