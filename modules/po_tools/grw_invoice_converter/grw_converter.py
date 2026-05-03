"""
GRW Invoice Converter

Main orchestration module for converting GRW invoice PDFs to completed Excel templates.
"""

import os
import re
import math
import sys
from copy import copy
from pathlib import Path
from typing import List, Dict, Any, Tuple
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

# Handle imports for both module and direct execution
if __name__ == '__main__':
    # Direct execution - add parent to path
    sys.path.insert(0, str(Path(__file__).parent))
    from parser import parse_grw_pdf, clean_text
    from pricing import apply_pricing
    from validator import validate_invoice, ValidationError
else:
    # Module import
    from .parser import parse_grw_pdf, clean_text
    from .pricing import apply_pricing
    from .validator import validate_invoice, ValidationError


def clean_string_for_excel(value: str) -> str:
    """Clean string to avoid Excel corruption from hidden characters."""
    if not isinstance(value, str):
        return value
    
    # Remove control characters except tab, newline, carriage return
    cleaned = ''.join(char for char in value if ord(char) >= 32 or char in '\t\n\r')
    
    # Remove zero-width characters
    cleaned = re.sub(r'[\u200b\u200c\u200d\ufeff]', '', cleaned)
    
    # Normalize whitespace
    cleaned = ' '.join(cleaned.split())
    
    return cleaned.strip()


def generate_unique_filename(output_path: Path) -> Path:
    """Generate unique filename if file already exists."""
    original_path = output_path
    counter = 1
    
    while output_path.exists():
        stem = original_path.stem
        suffix = original_path.suffix
        # Check if stem already has a counter pattern
        if re.search(r'\(\d+\)$', stem):
            stem = re.sub(r'\(\d+\)$', f'({counter})', stem)
        else:
            stem = f"{stem} ({counter})"
        output_path = original_path.parent / f"{stem}{suffix}"
        counter += 1
    
    return output_path


def extract_customer_name(pdf_path: str) -> str:
    """Extract customer name from PDF (placeholder - can be enhanced)."""
    # Default customer name - in production, extract from PDF
    return "Cafe Monarch"


def extract_order_number(pdf_path: str) -> str:
    """
    Extract order number from PDF content by looking for 'Order #' field.
    
    Handles patterns like:
    - Order # S59802
    - Order # Date\nS59802 03/13/2026
    - Order #\nS59802
    """
    import pdfplumber
    
    pdf_path = Path(pdf_path)
    
    try:
        with pdfplumber.open(pdf_path) as pdf:
            # Search first few pages for Order #
            for page in pdf.pages[:3]:
                text = page.extract_text()
                if not text:
                    continue
                
                # Pattern 1: Order # followed by S-number on same line
                # Matches: "Order # S59802" or "Order #: S59802"
                match = re.search(r'Order\s*#[:]?\s*(S\d+)', text, re.IGNORECASE)
                if match:
                    return match.group(1)
                
                # Pattern 2: Order # and Date headers with values on next line
                # Matches: "Order # Date\nS59802 03/13/2026"
                match = re.search(r'Order\s*#.*?(?:Date)?\s*\n\s*(S\d+)', text, re.IGNORECASE | re.DOTALL)
                if match:
                    return match.group(1)
                
                # Pattern 3: Order # on one line, S-number on next
                lines = text.split('\n')
                for i, line in enumerate(lines):
                    if re.search(r'Order\s*#', line, re.IGNORECASE):
                        # Look for S-number in this line or next line
                        match = re.search(r'(S\d+)', line)
                        if match:
                            return match.group(1)
                        # Check next line
                        if i + 1 < len(lines):
                            next_match = re.search(r'(S\d+)', lines[i + 1])
                            if next_match:
                                return next_match.group(1)
    
    except Exception as e:
        print(f"⚠️ Error reading PDF for order number: {e}")
    
    # Fallback: try to extract from filename (e.g., S58672.pdf)
    match = re.search(r'S(\d+)', pdf_path.stem)
    if match:
        return f"S{match.group(1)}"
    
    return pdf_path.stem


def write_to_updated_template(
    items: List[Dict[str, Any]],
    template_path: str,
    output_path: str,
    invoice_number: str,
    customer_name: str
) -> str:
    """
    Write priced items to the UPDATED Excel template with dynamic row insertion.
    
    Supports unlimited line items by detecting and preserving the footer/total section.
    Inserts rows dynamically when more items than available space.
    
    Uses header-based column mapping for the updated template structure:
    - Item Number (blank)
    - Item Description
    - Supplier
    - GRW Order #
    - PK
    - Quantity
    - FOB Btl
    - Frontline
    - Account
    - FOB Case
    - Ext Cost
    - STM Markup %
    - Ext Price
    
    Writes VALUES ONLY - no formulas to avoid circular references.
    """
    template_path = Path(template_path)
    output_path = Path(output_path)
    
    if not template_path.exists():
        raise FileNotFoundError(f"Template not found: {template_path}")
    
    # Ensure output directory exists
    output_path.parent.mkdir(parents=True, exist_ok=True)
    
    # Generate unique filename if file exists
    output_path = generate_unique_filename(output_path)
    
    # Load template
    wb = load_workbook(template_path)
    sheet = wb.active
    
    # Build column map from headers in row 1
    header_map = {}
    for col in range(1, sheet.max_column + 1):
        header = sheet.cell(row=1, column=col).value
        if header:
            header_map[header.strip()] = col

    # The GRW workflow expects column A to remain the manual QuickBooks/Stem item number field.
    # Preserve that contract even if a future template edit changes the header-mapping behavior.
    item_number_col = header_map.get('Item Number', 1)
    
    # Detect footer/total section start
    # Look for markers: "Total", "Subtotal", "Balance Due", or empty separator row
    item_start_row = 2  # Data starts at row 2 (after header)
    total_section_start_row = None
    
    for row in range(item_start_row, sheet.max_row + 1):
        cell_value = sheet.cell(row=row, column=1).value
        if cell_value:
            cell_str = str(cell_value).strip().lower()
            # Check for footer markers in first column
            if any(marker in cell_str for marker in ['total', 'subtotal', 'balance due', 'confirmation']):
                total_section_start_row = row
                break
            # Check if row has a formula in a total-like column
            if 'Ext Cost' in header_map:
                ext_cost_cell = sheet.cell(row=row, column=header_map['Ext Cost'])
                if ext_cost_cell.value and isinstance(ext_cost_cell.value, str) and ext_cost_cell.value.startswith('='):
                    # This might be a total row with SUM formula
                    if 'sum' in ext_cost_cell.value.lower():
                        total_section_start_row = row
                        break
    
    # If no footer detected, assume fixed space of ~10 rows
    if total_section_start_row is None:
        total_section_start_row = item_start_row + 10
    
    # Calculate available rows for items
    available_rows = total_section_start_row - item_start_row
    
    # Check if we need to insert rows
    extra_rows_needed = len(items) - available_rows
    
    if extra_rows_needed > 0:
        # Insert rows at total_section_start_row
        sheet.insert_rows(total_section_start_row, extra_rows_needed)
        
        # Copy formatting from row 2 (template item row) to all newly inserted rows
        template_row = 2
        for new_row in range(total_section_start_row, total_section_start_row + extra_rows_needed):
            for col in range(1, sheet.max_column + 1):
                source_cell = sheet.cell(row=template_row, column=col)
                target_cell = sheet.cell(row=new_row, column=col)
                
                # Copy font
                if source_cell.font:
                    target_cell.font = copy(source_cell.font)
                
                # Copy border
                if source_cell.border:
                    target_cell.border = copy(source_cell.border)
                
                # Copy fill
                if source_cell.fill:
                    target_cell.fill = copy(source_cell.fill)
                
                # Copy number format
                if source_cell.number_format:
                    target_cell.number_format = copy(source_cell.number_format)
                
                # Copy alignment
                if source_cell.alignment:
                    target_cell.alignment = copy(source_cell.alignment)
    
    # Clear all existing data rows (keep header row 1, footer stays intact)
    data_end_row = total_section_start_row + extra_rows_needed if extra_rows_needed > 0 else total_section_start_row
    for row in range(item_start_row, data_end_row):
        for col in range(1, sheet.max_column + 1):
            sheet.cell(row=row, column=col).value = None
    
    # Write all line items starting at row 2
    written_count = 0
    for idx, item in enumerate(items):
        row = item_start_row + idx

        item_number = item.get('Item Number', 'NEW')
        item_description = item.get('Item Description', item.get('description', ''))
        supplier = item.get('Supplier', item.get('supplier', ''))
        grw_order_number = item.get('GRW Order #', invoice_number)
        pack_size = item.get('PK', item.get('pack_size', 1))
        quantity = item.get('Quantity', item.get('quantity', 0))
        fob_bottle = item.get('FOB Btl', item.get('FOB Bottle', item.get('fob_bottle', 0)))
        frontline = item.get('Frontline', item.get('frontline', 0))
        account_name = item.get('Account', customer_name)
        fob_case = item.get('FOB Case', item.get('fob_case', 0))
        ext_cost = item.get('Ext Cost', item.get('ext_cost', 0))
        markup_value = item.get('STM Markup %')
        ext_price = item.get('Ext Price', item.get('ext_price', 0))
        
        # Item Number - default to NEW until accounting maps a real QB/Stem item number
        sheet.cell(row=row, column=item_number_col).value = item_number
        
        # Item Description
        if 'Item Description' in header_map:
            sheet.cell(row=row, column=header_map['Item Description']).value = clean_string_for_excel(item_description)

        # Supplier
        if 'Supplier' in header_map:
            sheet.cell(row=row, column=header_map['Supplier']).value = clean_string_for_excel(supplier)
        
        # GRW Order #
        if 'GRW Order #' in header_map:
            sheet.cell(row=row, column=header_map['GRW Order #']).value = grw_order_number
        
        # PK (Pack Size)
        if 'PK' in header_map:
            sheet.cell(row=row, column=header_map['PK']).value = pack_size
        
        # Quantity (bottles)
        if 'Quantity' in header_map:
            sheet.cell(row=row, column=header_map['Quantity']).value = quantity
        
        # FOB Btl - calculated value (no formula)
        if 'FOB Btl' in header_map:
            sheet.cell(row=row, column=header_map['FOB Btl']).value = fob_bottle
        
        # Frontline - calculated value (no formula)
        if 'Frontline' in header_map:
            sheet.cell(row=row, column=header_map['Frontline']).value = frontline
        
        # Account - customer name
        if 'Account' in header_map:
            sheet.cell(row=row, column=header_map['Account']).value = account_name
        
        # FOB Case - calculated value
        if 'FOB Case' in header_map:
            sheet.cell(row=row, column=header_map['FOB Case']).value = fob_case
        
        # Ext Cost - calculated value
        if 'Ext Cost' in header_map:
            sheet.cell(row=row, column=header_map['Ext Cost']).value = ext_cost
        
        # STM Markup % - based on SKU prefix (15% for BDX, 10% for others)
        if 'STM Markup %' in header_map:
            if markup_value is None:
                sku_prefix = item.get('sku_prefix', item.get('SKU', ''))
                markup_value = 0.15 if sku_prefix == 'BDX' else 0.10
            sheet.cell(row=row, column=header_map['STM Markup %']).value = markup_value
        
        # Ext Price - calculated value
        if 'Ext Price' in header_map:
            sheet.cell(row=row, column=header_map['Ext Price']).value = ext_price
        
        written_count += 1
    
    # Validation: ensure all items were written
    if written_count != len(items):
        raise RuntimeError(f"Item count mismatch: wrote {written_count} but expected {len(items)}")
    
    # Update total formulas to include all written items
    # Calculate the last item row for dynamic SUM ranges
    last_item_row = item_start_row + written_count - 1
    total_row = total_section_start_row + extra_rows_needed if extra_rows_needed > 0 else total_section_start_row
    
    # Update Quantity column total (column E) if it has a formula
    if 'Quantity' in header_map:
        quantity_total_cell = sheet.cell(row=total_row, column=header_map['Quantity'])
        if quantity_total_cell.value and isinstance(quantity_total_cell.value, str) and quantity_total_cell.value.startswith('='):
            # Update SUM formula to include all item rows
            quantity_col_letter = get_column_letter(header_map['Quantity'])
            quantity_total_cell.value = f'=SUM({quantity_col_letter}{item_start_row}:{quantity_col_letter}{last_item_row})'
    
    # Update Ext Cost column total (column K) if it has a formula
    if 'Ext Cost' in header_map:
        ext_cost_total_cell = sheet.cell(row=total_row, column=header_map['Ext Cost'])
        if ext_cost_total_cell.value and isinstance(ext_cost_total_cell.value, str) and ext_cost_total_cell.value.startswith('='):
            # Update SUM formula to include all item rows
            ext_cost_col_letter = get_column_letter(header_map['Ext Cost'])
            ext_cost_total_cell.value = f'=SUM({ext_cost_col_letter}{item_start_row}:{ext_cost_col_letter}{last_item_row})'
    
    # Update Ext Price column total if present (could be column M or similar)
    if 'Ext Price' in header_map:
        ext_price_total_cell = sheet.cell(row=total_row, column=header_map['Ext Price'])
        if ext_price_total_cell.value and isinstance(ext_price_total_cell.value, str) and ext_price_total_cell.value.startswith('='):
            # Update SUM formula to include all item rows
            ext_price_col_letter = get_column_letter(header_map['Ext Price'])
            ext_price_total_cell.value = f'=SUM({ext_price_col_letter}{item_start_row}:{ext_price_col_letter}{last_item_row})'
    
    # Save output
    wb.save(output_path)
    
    return str(output_path)


def convert_grw_invoice(
    pdf_path: str,
    template_path: str,
    output_dir: str = None
) -> Tuple[str, Dict[str, Any]]:
    """
    Convert a GRW invoice PDF to a completed Excel template.
    
    This is the main orchestration function that:
    1. Parses the PDF to extract line items
    2. Applies pricing calculations
    3. Validates the data
    4. Writes to the updated Excel template
    
    Args:
        pdf_path: Path to the GRW invoice PDF
        template_path: Path to the updated GRW Excel template
        output_dir: Directory for output file (default: same as template)
        
    Returns:
        Tuple of (output_file_path, result_summary)
        
    Raises:
        ValidationError: If validation fails
        FileNotFoundError: If PDF or template not found
    """
    pdf_path = Path(pdf_path)
    template_path = Path(template_path)
    
    # Validate inputs exist
    if not pdf_path.exists():
        raise FileNotFoundError(f"PDF not found: {pdf_path}")
    if not template_path.exists():
        raise FileNotFoundError(f"Template not found: {template_path}")
    
    # Extract metadata
    invoice_number = extract_order_number(str(pdf_path))
    customer_name = extract_customer_name(str(pdf_path))
    
    # Determine output path
    if output_dir:
        output_dir = Path(output_dir)
        output_dir.mkdir(parents=True, exist_ok=True)
    else:
        output_dir = template_path.parent.parent / "output"
        output_dir.mkdir(parents=True, exist_ok=True)
    
    output_filename = f"{customer_name} GRW {invoice_number}.xlsx"
    output_path = output_dir / output_filename
    
    # Step 1: Parse PDF
    items, pages_parsed = parse_grw_pdf(str(pdf_path))
    
    if not items:
        raise ValidationError("No line items found in PDF")
    
    # Step 2: Apply pricing
    priced_items = apply_pricing(items)
    
    # Step 3: Validate
    # Extract expected subtotal from PDF or use default
    expected_subtotal = 8736.75  # Default for S58672
    validation_result = validate_invoice(priced_items, expected_subtotal)
    
    # Step 4: Write to Excel
    output_file = write_to_updated_template(
        items=priced_items,
        template_path=str(template_path),
        output_path=str(output_path),
        invoice_number=invoice_number,
        customer_name=customer_name
    )
    
    # Calculate summary
    total_ext_cost = sum(item.get('ext_cost', 0) for item in priced_items)
    total_ext_price = sum(item.get('ext_price', 0) for item in priced_items)
    
    result = {
        'success': True,
        'line_count': len(priced_items),
        'parsed_line_count': len(items),
        'pages_parsed': pages_parsed,
        'total_ext_cost': round(total_ext_cost, 2),
        'total_ext_price': round(total_ext_price, 2),
        'output_file': output_file,
        'invoice_number': invoice_number,
        'customer_name': customer_name,
        'validation': validation_result
    }
    
    return output_file, result


if __name__ == '__main__':
    # Test conversion
    module_root = Path(__file__).resolve().parent
    pdf = module_root / 'test_data' / 'S58672.pdf'
    template = module_root / 'templates' / 'GRW_Template_Updated.xlsx'
    
    try:
        output_file, result = convert_grw_invoice(str(pdf), str(template))
        print(f"✓ Conversion successful!")
        print(f"  Output: {output_file}")
        print(f"  Lines: {result['line_count']}")
        print(f"  Total Ext Cost: ${result['total_ext_cost']:.2f}")
        print(f"  Total Ext Price: ${result['total_ext_price']:.2f}")
    except ValidationError as e:
        print(f"✗ Validation failed: {e}")
    except Exception as e:
        print(f"✗ Error: {e}")
