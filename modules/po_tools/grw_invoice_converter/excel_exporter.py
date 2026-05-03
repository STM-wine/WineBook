"""
GRW Invoice Excel Exporter

Writes priced invoice data to Excel template with values only.
"""

import re
from pathlib import Path
from typing import List, Dict, Any
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter


def clean_string_for_excel(value: str) -> str:
    """Clean string to avoid Excel corruption from hidden characters."""
    if not isinstance(value, str):
        return value
    
    # Remove control characters except tab, newline, carriage return
    cleaned = ''.join(char for char in value if ord(char) >= 32 or char in '\t\n\r')
    
    # Remove zero-width characters
    cleaned = re.sub(r'[\u200b\u200c\u200d\ufeff]', '', cleaned)
    
    # Normalize whitespace
    cleaned = ' '.join(cleaned.split())
    
    return cleaned.strip()


def export_to_excel(
    items: List[Dict[str, Any]],
    template_path: str,
    output_path: str,
    invoice_number: str = ""
) -> str:
    """
    Export priced items to Excel template.
    
    Writes VALUES ONLY (no formulas) to avoid circular references.
    
    Args:
        items: List of priced line items
        template_path: Path to GRW_Template.xlsx
        output_path: Path for output file
        invoice_number: Invoice number for metadata
        
    Returns:
        Path to created output file
    """
    template_path = Path(template_path)
    output_path = Path(output_path)
    
    if not template_path.exists():
        raise FileNotFoundError(f"Template not found: {template_path}")
    
    # Ensure output directory exists
    output_path.parent.mkdir(parents=True, exist_ok=True)
    
    # Check for file collision and generate unique filename if needed
    original_path = output_path
    counter = 1
    while output_path.exists():
        stem = original_path.stem
        suffix = original_path.suffix
        # Check if stem already has a counter pattern
        if re.search(r'\(\d+\)$', stem):
            stem = re.sub(r'\(\d+\)$', f'({counter})', stem)
        else:
            stem = f"{stem} ({counter})"
        output_path = original_path.parent / f"{stem}{suffix}"
        counter += 1
    
    # Load template
    wb = load_workbook(template_path)
    sheet = wb.active
    
    # Clear existing data rows (keep header row 1)
    # Find the last row with data
    max_row = sheet.max_row
    if max_row > 1:
        # Clear rows 2 onwards to remove template example data
        for row in range(2, max_row + 1):
            for col in range(1, sheet.max_column + 1):
                sheet.cell(row=row, column=col).value = None
    
    # Build column map from headers in row 1
    header_map = {}
    for col in range(1, sheet.max_column + 1):
        header = sheet.cell(row=1, column=col).value
        if header:
            header_map[header.strip()] = col
    
    # Write data starting at row 2
    for idx, item in enumerate(items, start=2):
        # A: Item Number - leave blank
        sheet.cell(row=idx, column=1).value = None
        
        # B: Item Description
        sheet.cell(row=idx, column=2).value = clean_string_for_excel(item.get('description', ''))
        
        # Map by exact header names
        # GRW Order # (G) → invoice_number
        if 'GRW Order #' in header_map:
            sheet.cell(row=idx, column=header_map['GRW Order #']).value = invoice_number
        
        # PK (H) → pack_size
        if 'PK' in header_map:
            sheet.cell(row=idx, column=header_map['PK']).value = item.get('pack_size', 1)
        
        # Quantity (I) → quantity
        if 'Quantity' in header_map:
            sheet.cell(row=idx, column=header_map['Quantity']).value = item.get('quantity', 0)
        
        # FOB Btl (J) → fob_bottle
        if 'FOB Btl' in header_map:
            sheet.cell(row=idx, column=header_map['FOB Btl']).value = item.get('fob_bottle', 0)
        
        # Frontline (K) → frontline
        if 'Frontline' in header_map:
            sheet.cell(row=idx, column=header_map['Frontline']).value = item.get('frontline', 0)
        
        # Account (L) → frontline (same as Frontline)
        if 'Account' in header_map:
            sheet.cell(row=idx, column=header_map['Account']).value = item.get('frontline', 0)
        
        # FOB Case (N) → fob_case
        if 'FOB Case' in header_map:
            sheet.cell(row=idx, column=header_map['FOB Case']).value = item.get('fob_case', 0)
        
        # Ext Cost (O) → ext_cost
        if 'Ext Cost' in header_map:
            sheet.cell(row=idx, column=header_map['Ext Cost']).value = item.get('ext_cost', 0)
    
    # Copy account structure from row 2 to all data rows
    # This preserves the COGS/Asset/Income account references
    if max_row >= 2 and len(items) > 0:
        template_row = 2  # Row with template structure
        for data_row in range(2, 2 + len(items)):
            for col in [3, 4, 5, 6]:  # C, D, E, F - account columns
                template_value = sheet.cell(row=template_row, column=col).value
                if template_value and data_row != template_row:
                    sheet.cell(row=data_row, column=col).value = template_value
    
    # Save output
    wb.save(output_path)
    
    return str(output_path)


def export_with_vintage_column(
    items: List[Dict[str, Any]],
    template_path: str,
    output_path: str,
    invoice_number: str = ""
) -> str:
    """
    Export with vintage and size columns if template supports it.
    
    Some templates have separate columns for vintage and size.
    This version handles extended column layouts.
    """
    template_path = Path(template_path)
    output_path = Path(output_path)
    
    if not template_path.exists():
        raise FileNotFoundError(f"Template not found: {template_path}")
    
    output_path.parent.mkdir(parents=True, exist_ok=True)
    
    wb = load_workbook(template_path)
    sheet = wb.active
    
    # Clear data rows
    max_row = sheet.max_row
    if max_row > 1:
        for row in range(2, max_row + 1):
            for col in range(1, sheet.max_column + 1):
                sheet.cell(row=row, column=col).value = None
    
    # Write data
    for idx, item in enumerate(items, start=2):
        col_idx = 1
        
        # Item Number - blank
        sheet.cell(row=idx, column=col_idx).value = None
        col_idx += 1
        
        # Description (use new formatted description)
        sheet.cell(row=idx, column=col_idx).value = clean_string_for_excel(item.get('description', ''))
        col_idx += 1
        
        # Write vintage if there's space (check if column exists)
        if sheet.max_column >= col_idx:
            sheet.cell(row=idx, column=col_idx).value = item.get('vintage', '')
            col_idx += 1
        
        # Write size if there's space
        if sheet.max_column >= col_idx:
            sheet.cell(row=idx, column=col_idx).value = item.get('size', '750mL')
            col_idx += 1
        
        # Continue with remaining columns
        remaining_cols = ['', '', invoice_number,  # COGS, Asset, Income, PO
                         item.get('quantity', 0),
                         item.get('fob_bottle', 0),
                         item.get('fob_case', 0),
                         item.get('pack_size', 1),
                         item.get('frontline', 0),
                         item.get('ext_cost', 0),
                         item.get('ext_price', 0)]
        
        for value in remaining_cols:
            if sheet.max_column >= col_idx:
                sheet.cell(row=idx, column=col_idx).value = value
                col_idx += 1
    
    wb.save(output_path)
    return str(output_path)


if __name__ == '__main__':
    # Test export
    test_items = [
        {
            'sku_prefix': 'BDX',
            'clean_description': 'Chateau Haut Brion',
            'vintage': 1998,
            'size': '750mL',
            'unit_price': 695.00,
            'pack_size': 1,
            'ordered_qty': 1,
            'quantity': 1,
            'fob_bottle': 695.00,
            'fob_case': 695.00,
            'frontline': 800,
            'ext_cost': 695.00,
            'ext_price': 800.00,
        },
        {
            'sku_prefix': 'USR',
            'clean_description': 'MacDonald Vineyards To-Kalon Cabernet',
            'vintage': 2022,
            'size': '750mL',
            'unit_price': 2199.75,
            'pack_size': 3,
            'ordered_qty': 1,
            'quantity': 3,
            'fob_bottle': 733.25,
            'fob_case': 2199.75,
            'frontline': 803,
            'ext_cost': 2199.75,
            'ext_price': 2409.00,
        }
    ]
    
    module_root = Path(__file__).resolve().parent
    template = module_root / 'templates' / 'GRW_Template.xlsx'
    output = module_root / 'output' / 'test_export.xlsx'
    
    result = export_to_excel(test_items, str(template), str(output), 'S58672')
    print(f"Exported to: {result}")
