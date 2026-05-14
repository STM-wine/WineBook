import unittest
from tempfile import TemporaryDirectory
from io import BytesIO

from openpyxl import Workbook, load_workbook
import pandas as pd

from scripts.process_daily_vinosmith_email import (
    AttachmentCandidate,
    classify_attachments,
    dedupe_attachments,
    mailbox_search_targets,
    safe_filename,
    storage_path,
)
from stem_order.core import calculate_reorder_recommendations, normalize_planning_sku
from stem_order.ingest import (
    clean_importer_name,
    load_importers_csv,
    map_rads_columns,
    map_rb6_columns,
    normalize_columns,
    supplier_logistics_rows_to_frame,
)
from stem_order.pipeline import format_display_dataframe, select_raw_output
from stem_order.dashboard import (
    active_po_draft_message,
    approval_metrics,
    buyer_updates_from_editor,
    buyer_workbench_dataframe,
    california_truck_summary,
    dashboard_metrics,
    filter_recommendations,
    format_dashboard_dataframe,
    importer_groups,
    importer_workbench_summary,
    importer_workflow_status,
    location_summary,
    po_draft_lines_dataframe,
    po_drafts_dataframe,
    po_export_dataframe,
    po_template_xlsx_bytes,
    recalculate_working_recommendation,
    recommendations_to_dataframe,
    risk_counts,
)
from stem_order.supabase_repository import SupabaseRepository


class IngestTests(unittest.TestCase):
    def test_normalize_columns_handles_duplicates_and_punctuation(self):
        df = pd.DataFrame(columns=[" Wine: External ID (1) ", "Available Inventory", "Available Inventory"])

        normalized = normalize_columns(df)

        self.assertEqual(
            list(normalized.columns),
            ["wine_external_id_1", "available_inventory", "available_inventory_1"],
        )

    def test_column_mapping_matches_mark_exports(self):
        rb6 = normalize_columns(
            pd.DataFrame(
                columns=[
                    "Name",
                    "Code",
                    "Available Inventory",
                    "On Order",
                    "Importer",
                    "FOB",
                ]
            )
        )
        rads = normalize_columns(
            pd.DataFrame(
                columns=[
                    "Account Name",
                    "Wine Name",
                    "Quantity",
                    "Date (mm/dd/yyyy)",
                    "Product Code",
                ]
            )
        )

        self.assertEqual(map_rb6_columns(rb6)["description"], "name")
        self.assertEqual(map_rb6_columns(rb6)["available_inventory"], "available_inventory")
        self.assertEqual(map_rb6_columns(rb6)["on_order"], "on_order")
        self.assertEqual(map_rads_columns(rads)["product_name"], "wine_name")
        self.assertEqual(map_rads_columns(rads)["quantity"], "quantity")
        self.assertEqual(map_rads_columns(rads)["date"], "date_mm_dd_yyyy")

    def test_rb6_on_order_prefers_true_on_order_header(self):
        columns = [
            "Name",
            "Importer",
            "Available Inventory",
            'Estimated # of Intervals Supply Remaining with "On Order" Considered',
            "On Order",
        ]
        rb6 = normalize_columns(pd.DataFrame(columns=columns))

        self.assertEqual(map_rb6_columns(rb6)["on_order"], "on_order")

    def test_clean_importer_name_collapses_whitespace(self):
        self.assertEqual(clean_importer_name("  Barnard   Griffin "), "barnard griffin")

    def test_missing_importers_csv_degrades_to_empty_frame(self):
        data, loaded, warning = load_importers_csv("/tmp/does-not-exist/importers.csv")

        self.assertFalse(loaded)
        self.assertIn("not found", warning)
        self.assertIn("eta_days", data.columns)

    def test_importers_csv_accepts_laid_in_per_bottle_alias(self):
        with TemporaryDirectory() as temp_dir:
            path = f"{temp_dir}/importers.csv"
            pd.DataFrame(
                [
                    {
                        "name": "Supplier A",
                        "eta_days": 14,
                        "laid_in_per_bottle": 1.25,
                    }
                ]
            ).to_csv(path, index=False)

            data, loaded, warning = load_importers_csv(path)

        self.assertTrue(loaded)
        self.assertIsNone(warning)
        self.assertIn("trucking_cost_per_bottle", data.columns)
        self.assertEqual(data.loc[0, "trucking_cost_per_bottle"], 1.25)

    def test_supplier_rows_shape_like_importer_logistics(self):
        data = supplier_logistics_rows_to_frame(
            [
                {
                    "name": "Supplier A",
                    "eta_days": 14,
                    "pick_up_location": "California",
                    "trucking_cost_per_bottle": 1.25,
                    "active": True,
                }
            ]
        )

        self.assertEqual(data.loc[0, "importer_name"], "Supplier A")
        self.assertEqual(data.loc[0, "importer_name_clean"], "supplier a")
        self.assertEqual(data.loc[0, "trucking_cost_per_bottle"], 1.25)


class CalculatorTests(unittest.TestCase):
    def _trend_for_sales(self, sales_rows):
        rb6 = pd.DataFrame(
            [
                {
                    "name": "Trend Wine 2025 12/750ml",
                    "available_inventory": 0,
                    "on_order": 0,
                    "fob": 10,
                    "pack_size": 12,
                }
            ]
        )
        sales = pd.DataFrame(sales_rows)
        return calculate_reorder_recommendations(rb6, sales).iloc[0]

    def test_normalize_planning_sku_removes_vintage_but_keeps_pack_size(self):
        self.assertEqual(
            normalize_planning_sku("Pavette Sauvignon Blanc 2025 12/750ml"),
            "pavette sauvignon blanc 12/750ml",
        )

    def test_btg_and_core_flags_drive_target_days(self):
        rb6 = pd.DataFrame(
            [
                {
                    "name": "BTG Wine 2025 12/750ml",
                    "available_inventory": 0,
                    "on_order": 0,
                    "fob": 10,
                    "pack_size": 12,
                    "is_btg": "Yes",
                    "is_core": "No",
                },
                {
                    "name": "Core Wine 2025 12/750ml",
                    "available_inventory": 0,
                    "on_order": 0,
                    "fob": 10,
                    "pack_size": 12,
                    "is_btg": "No",
                    "is_core": "Yes",
                },
                {
                    "name": "Standard Wine 2025 12/750ml",
                    "available_inventory": 0,
                    "on_order": 0,
                    "fob": 10,
                    "pack_size": 12,
                    "is_btg": "No",
                    "is_core": "No",
                },
            ]
        )
        sales = pd.DataFrame(
            [
                {"wine_name": "BTG Wine 2024 12/750ml", "quantity": 30, "date": "2026-04-01"},
                {"wine_name": "Core Wine 2024 12/750ml", "quantity": 30, "date": "2026-04-01"},
                {"wine_name": "Standard Wine 2024 12/750ml", "quantity": 30, "date": "2026-04-01"},
            ]
        )

        result = calculate_reorder_recommendations(rb6, sales).set_index("Name")

        self.assertEqual(result.loc["BTG Wine 2025 12/750ml", "target_days"], 45)
        self.assertEqual(result.loc["Core Wine 2025 12/750ml", "target_days"], 30)
        self.assertEqual(result.loc["Standard Wine 2025 12/750ml", "target_days"], 30)
        self.assertEqual(result.loc["BTG Wine 2025 12/750ml", "recommendation_status"], "rejected")
        self.assertEqual(result.loc["BTG Wine 2025 12/750ml", "approved_qty"], 0)

    def test_true_available_subtracts_unconfirmed_line_item_quantity(self):
        rb6 = pd.DataFrame(
            [
                {
                    "name": "Allocated Wine 2025 12/750ml",
                    "available_inventory": 10,
                    "unconfirmed_line_item_qty": 3,
                    "on_order": 0,
                    "fob": 10,
                    "pack_size": 12,
                },
                {
                    "name": "Oversold Wine 2025 12/750ml",
                    "available_inventory": 2,
                    "unconfirmed_line_item_qty": 5,
                    "on_order": 0,
                    "fob": 10,
                    "pack_size": 12,
                },
            ]
        )
        sales = pd.DataFrame(
            [
                {"wine_name": "Allocated Wine 2024 12/750ml", "quantity": 12, "date": "2026-04-01"},
                {"wine_name": "Oversold Wine 2024 12/750ml", "quantity": 12, "date": "2026-04-01"},
            ]
        )

        result = calculate_reorder_recommendations(rb6, sales).set_index("Name")

        self.assertEqual(result.loc["Allocated Wine 2025 12/750ml", "true_available"], 7)
        self.assertEqual(result.loc["Oversold Wine 2025 12/750ml", "true_available"], 0)

    def test_sales_windows_forecasts_and_velocity_trend_are_calculated(self):
        rb6 = pd.DataFrame(
            [
                {
                    "name": "Trend Wine 2025 12/750ml",
                    "available_inventory": 0,
                    "on_order": 0,
                    "fob": 10,
                    "pack_size": 12,
                }
            ]
        )
        sales = pd.DataFrame(
            [
                {"wine_name": "Trend Wine 2024 12/750ml", "quantity": 10, "date": "2026-04-01"},
                {"wine_name": "Trend Wine 2024 12/750ml", "quantity": 20, "date": "2026-03-10"},
                {"wine_name": "Trend Wine 2024 12/750ml", "quantity": 30, "date": "2026-02-01"},
                {"wine_name": "Trend Wine 2024 12/750ml", "quantity": 40, "date": "2025-05-15"},
            ]
        )

        result = calculate_reorder_recommendations(rb6, sales).iloc[0]

        self.assertEqual(result["last_30_day_sales"], 30)
        self.assertEqual(result["last_60_day_sales"], 60)
        self.assertEqual(result["last_90_day_sales"], 60)
        self.assertEqual(result["prior_30_day_sales"], 30)
        self.assertEqual(result["next_30_day_forecast"], 40)
        self.assertEqual(result["velocity_trend_pct"], 0)
        self.assertIn(result["risk_level"], ["High", "Medium", "Low", "No Sales", "Unknown"])

    def test_velocity_trend_positive_against_prior_30_days(self):
        result = self._trend_for_sales(
            [
                {"wine_name": "Trend Wine 2024 12/750ml", "quantity": 20, "date": "2026-04-01"},
                {"wine_name": "Trend Wine 2024 12/750ml", "quantity": 10, "date": "2026-02-15"},
            ]
        )

        self.assertEqual(result["last_30_day_sales"], 20)
        self.assertEqual(result["prior_30_day_sales"], 10)
        self.assertEqual(result["velocity_trend_pct"], 100)

    def test_velocity_trend_negative_against_prior_30_days(self):
        result = self._trend_for_sales(
            [
                {"wine_name": "Trend Wine 2024 12/750ml", "quantity": 5, "date": "2026-04-01"},
                {"wine_name": "Trend Wine 2024 12/750ml", "quantity": 10, "date": "2026-02-15"},
            ]
        )

        self.assertEqual(result["velocity_trend_pct"], -50)

    def test_velocity_trend_flat_against_prior_30_days(self):
        result = self._trend_for_sales(
            [
                {"wine_name": "Trend Wine 2024 12/750ml", "quantity": 10, "date": "2026-04-01"},
                {"wine_name": "Trend Wine 2024 12/750ml", "quantity": 10, "date": "2026-02-15"},
            ]
        )

        self.assertEqual(result["velocity_trend_pct"], 0)

    def test_velocity_trend_marks_new_when_prior_30_days_are_zero(self):
        result = self._trend_for_sales(
            [
                {"wine_name": "Trend Wine 2024 12/750ml", "quantity": 10, "date": "2026-04-01"},
            ]
        )

        self.assertEqual(result["prior_30_day_sales"], 0)
        self.assertTrue(pd.isna(result["velocity_trend_pct"]))
        self.assertEqual(result["velocity_trend_label"], "New")

    def test_pipeline_output_formatting_keeps_raw_numeric_values(self):
        recommendations = pd.DataFrame(
            [
                {
                    "planning_sku": "wine 12/750ml",
                    "Name": "Wine 2025 12/750ml",
                    "vintage": 2025.0,
                    "true_available": 1.0,
                    "weekly_velocity": 2.345,
                    "order_cost": 1234.5,
                    "recommended_qty_rounded": 12,
                }
            ]
        )

        raw_df = select_raw_output(recommendations)
        display_df = format_display_dataframe(raw_df)

        self.assertEqual(raw_df.loc[0, "order_cost"], 1234.5)
        self.assertEqual(display_df.loc[0, "order_cost"], "$1,234.50")
        self.assertEqual(display_df.loc[0, "vintage"], "2025")


class DashboardTests(unittest.TestCase):
    def test_all_order_review_display_sorts_and_uses_landed_costs(self):
        df = recommendations_to_dataframe(
            [
                {
                    "supplier_name": "Supplier B",
                    "product_name": "Zinfandel",
                    "product_code": "B2",
                    "planning_sku": "zinfandel",
                    "recommended_qty_rounded": 10,
                    "last_30_day_sales": 5,
                    "order_cost": 100.0,
                    "landed_cost": 125.0,
                    "trucking_cost_per_bottle": 2.5,
                },
                {
                    "supplier_name": "Supplier A",
                    "product_name": "Merlot",
                    "product_code": "A2",
                    "planning_sku": "merlot",
                    "recommended_qty_rounded": 6,
                    "last_30_day_sales": 8,
                    "order_cost": 72.0,
                    "landed_cost": 90.0,
                    "trucking_cost_per_bottle": 3.0,
                },
                {
                    "supplier_name": "Supplier A",
                    "product_name": "Cabernet",
                    "product_code": "A1",
                    "planning_sku": "cabernet",
                    "recommended_qty_rounded": 12,
                    "last_30_day_sales": 9,
                    "order_cost": 120.0,
                    "landed_cost": 138.0,
                    "trucking_cost_per_bottle": 1.5,
                },
            ]
        )

        display = format_dashboard_dataframe(df)

        self.assertEqual(display["Supplier"].tolist(), ["Supplier A", "Supplier A", "Supplier B"])
        self.assertEqual(display["Wine"].tolist(), ["Cabernet", "Merlot", "Zinfandel"])
        self.assertNotIn("Planning SKU", display.columns)
        self.assertNotIn("Recommended Cost", display.columns)
        self.assertIn("Total Wine Cost", display.columns)
        self.assertIn("Total Laid In Cost", display.columns)
        self.assertIn("Estimated Cost", display.columns)
        self.assertEqual(display.loc[0, "Total Wine Cost"], "$120.00")
        self.assertEqual(display.loc[0, "Total Laid In Cost"], "$18.00")
        self.assertEqual(display.loc[0, "Estimated Cost"], "$138.00")

    def test_dashboard_filters_and_po_export(self):
        df = recommendations_to_dataframe(
            [
                {
                    "supplier_name": "Supplier A",
                    "product_name": "Wine A",
                    "product_code": "A1",
                    "planning_sku": "wine a",
                    "reorder_status": "URGENT",
                    "recommendation_status": "rejected",
                    "approved_qty": 12,
                    "risk_level": "High",
                    "recommended_qty_rounded": 12,
                    "last_30_day_sales": 50,
                    "order_cost": 120.0,
                    "landed_cost": 132.0,
                    "pickup_location": "California",
                },
                {
                    "supplier_name": "Supplier B",
                    "product_name": "Wine B",
                    "product_code": "B1",
                    "planning_sku": "wine b",
                    "reorder_status": "OK",
                    "recommendation_status": "rejected",
                    "approved_qty": 0,
                    "risk_level": "Low",
                    "recommended_qty_rounded": 0,
                    "last_30_day_sales": 5,
                    "order_cost": 0.0,
                    "landed_cost": 0.0,
                    "pickup_location": "Washington",
                },
            ]
        )

        metrics = dashboard_metrics(df)
        filtered = filter_recommendations(df, supplier="Supplier A", statuses=["URGENT"])
        approved = filtered.copy()
        approved["recommendation_status"] = "approved"
        po_df = po_export_dataframe(approved)

        self.assertEqual(metrics.urgent_skus, 1)
        self.assertEqual(metrics.recommended_bottles, 12)
        self.assertEqual(len(filtered), 1)
        self.assertEqual(po_df.loc[0, "Quantity"], 12)

    def test_buyer_workbench_is_importer_scoped_and_decision_first(self):
        original = recommendations_to_dataframe(
            [
                {
                    "id": "rec-1",
                    "supplier_name": "Importer A",
                    "product_name": "Slower Wine",
                    "is_core": True,
                    "is_btg": False,
                    "true_available": 6,
                    "on_order": 6,
                    "last_30_day_sales": 12,
                    "last_60_day_sales": 24,
                    "last_90_day_sales": 36,
                    "last_365_day_sales": 144,
                    "next_30_day_forecast": 8,
                    "next_60_day_forecast": 16,
                    "next_90_day_forecast": 24,
                    "weekly_velocity": 3,
                    "velocity_trend_pct": -10,
                    "weeks_on_hand_with_on_order": 4,
                    "recommended_qty_rounded": 12,
                    "recommendation_status": "rejected",
                    "approved_qty": 0,
                    "order_cost": 120,
                    "fob": 10,
                    "pack_size": 12,
                },
                {
                    "id": "rec-2",
                    "supplier_name": "Importer A",
                    "product_name": "Faster Wine",
                    "is_core": False,
                    "is_btg": True,
                    "true_available": 0,
                    "on_order": 0,
                    "last_30_day_sales": 48,
                    "last_60_day_sales": 96,
                    "last_90_day_sales": 144,
                    "last_365_day_sales": 576,
                    "next_30_day_forecast": 40,
                    "next_60_day_forecast": 80,
                    "next_90_day_forecast": 120,
                    "weekly_velocity": 12,
                    "velocity_trend_pct": 15,
                    "weeks_on_hand_with_on_order": 0,
                    "recommended_qty_rounded": 48,
                    "recommendation_status": "approved",
                    "approved_qty": 48,
                    "order_cost": 480,
                    "fob": 10,
                    "pack_size": 12,
                },
            ]
        )

        compact = buyer_workbench_dataframe(original)
        expanded = buyer_workbench_dataframe(original, show_history=True, show_forecast=True)

        self.assertNotIn("Supplier", compact.columns)
        self.assertEqual(compact.loc[0, "Recommended Qty"], 48)
        self.assertTrue(compact.loc[0, "Wine"].startswith("#1 Faster Wine"))
        self.assertIn("🍷", compact.loc[0, "Wine"])
        self.assertIn("⭐", compact.loc[1, "Wine"])
        self.assertNotIn("60d Sales", compact.columns)
        self.assertIn("60d Sales", expanded.columns)
        self.assertIn("LY Next 90d Forecast", expanded.columns)
        self.assertEqual(compact.loc[0, "Velocity Trend"], "+15%")
        self.assertEqual(compact.loc[0, "Weekly Velocity"], 12)
        self.assertEqual(compact.loc[0, "Weeks w/ Recommended"], 4.0)

    def test_buyer_workbench_velocity_trend_displays_new_label(self):
        original = recommendations_to_dataframe(
            [
                {
                    "id": "rec-1",
                    "supplier_name": "Importer A",
                    "product_name": "New Mover",
                    "true_available": 0,
                    "on_order": 0,
                    "last_30_day_sales": 12,
                    "prior_30_day_sales": 0,
                    "weekly_velocity": 3,
                    "velocity_trend_pct": None,
                    "velocity_trend_label": "New",
                    "recommended_qty_rounded": 12,
                    "recommendation_status": "rejected",
                    "approved_qty": 0,
                }
            ]
        )

        compact = buyer_workbench_dataframe(original)

        self.assertEqual(compact.loc[0, "Velocity Trend"], "New")

    def test_buyer_workbench_recommended_qty_override_is_saved_when_approved(self):
        original = recommendations_to_dataframe(
            [
                {
                    "id": "rec-1",
                    "product_name": "Wine A",
                    "true_available": 6,
                    "on_order": 6,
                    "weekly_velocity": 3,
                    "weeks_on_hand_with_on_order": 4,
                    "recommended_qty_rounded": 12,
                    "recommendation_status": "rejected",
                    "approved_qty": 0,
                    "order_cost": 120,
                    "pack_size": 12,
                }
            ]
        )
        edited = buyer_workbench_dataframe(original)
        edited.loc[0, "Approval"] = True
        edited.loc[0, "Recommended Qty"] = 24

        updates = buyer_updates_from_editor(original, edited)

        self.assertEqual(
            updates,
            [{"id": "rec-1", "recommendation_status": "edited", "approved_qty": 24}],
        )

    def test_buyer_workbench_recommended_qty_can_change_after_approval(self):
        original = recommendations_to_dataframe(
            [
                {
                    "id": "rec-1",
                    "product_name": "Wine A",
                    "true_available": 6,
                    "on_order": 6,
                    "weekly_velocity": 3,
                    "recommended_qty_rounded": 12,
                    "recommendation_status": "edited",
                    "approved_qty": 24,
                    "order_cost": 120,
                }
            ]
        )
        edited = buyer_workbench_dataframe(original)
        edited.loc[0, "Recommended Qty"] = 36

        updates = buyer_updates_from_editor(original, edited)

        self.assertEqual(
            updates,
            [{"id": "rec-1", "recommendation_status": "edited", "approved_qty": 36}],
        )

    def test_buyer_workbench_approval_keeps_recommended_qty_when_weeks_unchanged(self):
        original = recommendations_to_dataframe(
            [
                {
                    "id": "rec-1",
                    "product_name": "Wine A",
                    "true_available": 3005,
                    "on_order": 0,
                    "weekly_velocity": 823.9356,
                    "weeks_on_hand_with_on_order": 3.65,
                    "recommended_qty_rounded": 2292,
                    "recommendation_status": "rejected",
                    "approved_qty": 0,
                    "order_cost": 120,
                    "pack_size": 12,
                }
            ]
        )
        edited = buyer_workbench_dataframe(original)
        edited.loc[0, "Approval"] = True

        updates = buyer_updates_from_editor(original, edited)

        self.assertEqual(
            updates,
            [{"id": "rec-1", "recommendation_status": "approved", "approved_qty": 2292}],
        )

    def test_recommended_weeks_recalculates_from_edited_quantity(self):
        display = pd.DataFrame(
            [
                {
                    "True Available": 10,
                    "On Order": 5,
                    "Recommended Qty": 30,
                    "Weekly Velocity": 5,
                    "Weeks w/ Recommended": 0,
                    "_FOB": 10,
                    "Est. Cost": 0,
                }
            ]
        )

        updated = recalculate_working_recommendation(display)

        self.assertEqual(updated.loc[0, "Weeks w/ Recommended"], 9.0)
        self.assertEqual(updated.loc[0, "Est. Cost"], 300)

    def test_approval_and_risk_metrics(self):
        df = recommendations_to_dataframe(
            [
                {
                    "supplier_name": "Supplier A",
                    "recommendation_status": "approved",
                    "approved_qty": 12,
                    "fob": 10,
                    "risk_level": "High",
                },
                {
                    "supplier_name": "Supplier A",
                    "recommendation_status": "edited",
                    "approved_qty": 6,
                    "fob": 8,
                    "risk_level": "Medium",
                },
                {
                    "supplier_name": "Supplier B",
                    "recommendation_status": "rejected",
                    "approved_qty": 0,
                    "fob": 9,
                    "risk_level": "Low",
                },
            ]
        )

        approvals = approval_metrics(df)
        risks = risk_counts(df)

        self.assertEqual(approvals.approved_lines, 2)
        self.assertEqual(approvals.approved_bottles, 18)
        self.assertEqual(approvals.approved_cost, 168)
        self.assertEqual(approvals.pending_lines, 1)
        self.assertEqual(risks, {"High": 1, "Medium": 1, "Low": 1})

    def test_importer_workbench_groups_by_value_and_status(self):
        df = recommendations_to_dataframe(
            [
                {
                    "supplier_name": "Importer B",
                    "product_name": "Wine B",
                    "recommended_qty_rounded": 12,
                    "recommendation_status": "rejected",
                    "approved_qty": 0,
                    "order_cost": 120,
                    "fob": 10,
                    "reorder_status": "URGENT",
                },
                {
                    "supplier_name": "Importer A",
                    "product_name": "Wine A",
                    "recommended_qty_rounded": 24,
                    "recommendation_status": "approved",
                    "approved_qty": 24,
                    "order_cost": 240,
                    "fob": 10,
                    "reorder_status": "LOW",
                },
                {
                    "supplier_name": "Importer A",
                    "product_name": "Wine C",
                    "recommended_qty_rounded": 12,
                    "recommendation_status": "edited",
                    "approved_qty": 12,
                    "order_cost": 120,
                    "fob": 10,
                    "reorder_status": "URGENT",
                },
            ]
        )

        summary = importer_workbench_summary(df)
        groups = importer_groups(df)

        self.assertEqual(summary.loc[0, "Importer"], "Importer A")
        self.assertEqual(summary.loc[0, "Status"], "Approved")
        self.assertEqual(summary.loc[1, "Status"], "Not Started")
        self.assertEqual(importer_workflow_status(df[df["supplier_name"] == "Importer A"], po_sent=True), "PO Sent")
        self.assertEqual(groups[0]["data"].iloc[0]["order_cost"], 240)

    def test_location_and_truck_summaries(self):
        df = recommendations_to_dataframe(
            [
                {
                    "supplier_name": "Supplier A",
                    "product_name": "Wine A",
                    "recommended_qty_rounded": 120,
                    "landed_cost": 1400,
                    "pickup_location": "California",
                }
            ]
        )

        summary = location_summary(df)
        truck = california_truck_summary(df)

        self.assertEqual(summary.loc[0, "Pickup Location"], "California")
        self.assertEqual(summary.loc[0, "Recommended Qty"], 120)
        self.assertEqual(truck["bottles_needed"], 10080)

    def test_po_draft_display_helpers_shape_exports(self):
        po_export = po_export_dataframe(
            recommendations_to_dataframe(
                [
                    {
                        "supplier_name": "Supplier A",
                        "product_name": "Wine A",
                        "product_code": "A1",
                        "planning_sku": "wine a",
                        "recommendation_status": "approved",
                        "approved_qty": 12,
                        "fob": 10,
                        "trucking_cost_per_bottle": 1.25,
                        "order_cost": 120,
                    }
                ]
            )
        )
        lines = po_draft_lines_dataframe(
            [
                {
                    "product_name": "Wine A",
                    "product_code": "A1",
                    "planning_sku": "wine a",
                    "approved_qty": 12,
                    "fob": 10,
                    "line_cost": 120,
                }
            ]
        )
        drafts = po_drafts_dataframe(
            [
                {
                    "id": "abcdef12-3456",
                    "supplier_name": "Supplier A",
                    "status": "ready_for_entry",
                    "created_at": "2026-05-06T10:00:00",
                    "notes": "Ready",
                }
            ]
        )

        self.assertEqual(po_export.loc[0, "Supplier"], "Supplier A")
        self.assertEqual(po_export.loc[0, "Quantity"], 12)
        self.assertEqual(po_export.loc[0, "FOB"], 10)
        self.assertEqual(po_export.loc[0, "Laid In Cost"], 1.25)
        self.assertNotIn("Planning SKU", po_export.columns)
        self.assertNotIn("Recommended Cost", po_export.columns)
        self.assertEqual(po_export.loc[0, "Total Wine Cost"], 120)
        self.assertEqual(po_export.loc[0, "Total Laid In Cost"], 15)
        self.assertEqual(po_export.loc[0, "Estimated Cost"], 135)
        self.assertEqual(lines.loc[0, "Quantity"], 12)
        self.assertEqual(lines.loc[0, "Estimated Cost"], 120)
        self.assertEqual(drafts.loc[0, "Draft ID"], "abcdef12")
        self.assertEqual(drafts.loc[0, "Status"], "Ready for Entry")

    def test_po_template_export_populates_mark_template_columns(self):
        po_df = pd.DataFrame(
            [
                {
                    "Supplier": "Supplier A",
                    "Wine": "Wine A 2024 12/750ml",
                    "Code": "A1",
                    "Planning SKU": "wine a",
                    "Quantity": 12,
                    "FOB": 10.5,
                    "Laid In Cost": 1.25,
                }
                ,
                {
                    "Supplier": "Supplier B",
                    "Wine": "Wine B 2024 12/750ml",
                    "Code": "B1",
                    "Quantity": 6,
                    "FOB": 12,
                    "Laid In Cost": 2,
                }
            ]
        )
        with TemporaryDirectory() as temp_dir:
            template_path = f"{temp_dir}/template.xlsx"
            workbook = Workbook()
            sheet = workbook.active
            sheet["A2"] = "Supplier/Importer"
            sheet["B2"] = "Producer"
            sheet["C2"] = "Quantity"
            sheet["D2"] = "Item Code"
            sheet["E2"] = "Item Description"
            sheet["F2"] = "FOB"
            sheet["G2"] = "Laid In Cost"
            workbook.save(template_path)

            output = po_template_xlsx_bytes(po_df, template_path)

        exported = load_workbook(BytesIO(output), data_only=True)
        sheet = exported.active
        self.assertEqual(sheet["A4"].value, "Supplier A")
        self.assertEqual(sheet["C4"].value, 12)
        self.assertEqual(sheet["D4"].value, "A1")
        self.assertEqual(sheet["E4"].value, "Wine A 2024 12/750ml")
        self.assertEqual(sheet["F4"].value, 10.5)
        self.assertEqual(sheet["G4"].value, 1.25)
        self.assertIsNone(sheet["A5"].value)
        self.assertEqual(sheet["A6"].value, "Supplier B")

    def test_active_po_draft_message_ignores_completed_and_cancelled_drafts(self):
        self.assertEqual(
            active_po_draft_message(
                [
                    {"id": "entered-1", "status": "entered_in_quickbooks"},
                    {"id": "cancelled-1", "status": "cancelled"},
                ]
            ),
            "",
        )
        self.assertIn(
            "abcdef12",
            active_po_draft_message([{"id": "abcdef12-3456", "status": "ready_for_entry"}]),
        )


class SupabaseRepositoryTests(unittest.TestCase):
    def test_purchase_order_status_validation(self):
        repo = SupabaseRepository(client=None)

        with self.assertRaises(ValueError):
            repo.update_purchase_order_draft_status("draft-1", "sent")

    def test_purchase_order_line_payload_uses_transitional_product_fields(self):
        repo = SupabaseRepository(client=None)

        payload = repo._purchase_order_line_payload(
            "draft-1",
            {
                "id": "recommendation-1",
                "product_name": "Wine A",
                "product_code": "A1",
                "planning_sku": "wine a",
                "recommended_qty_rounded": 12,
                "approved_qty": 6,
                "order_cost": 120.0,
                "fob": "10.0",
                "trucking_cost_per_bottle": 1.25,
                "diagnostics": {"fob": 10.0},
            },
        )

        self.assertEqual(payload["purchase_order_draft_id"], "draft-1")
        self.assertEqual(payload["recommendation_id"], "recommendation-1")
        self.assertEqual(payload["product_name"], "Wine A")
        self.assertEqual(payload["planning_sku"], "wine a")
        self.assertEqual(payload["recommended_qty"], 12)
        self.assertEqual(payload["approved_qty"], 6)
        self.assertEqual(payload["fob"], 10.0)
        self.assertEqual(payload["line_cost"], 60.0)
        self.assertEqual(payload["trucking_cost_per_bottle"], 1.25)
        self.assertEqual(payload["wine_cost"], 60.0)
        self.assertEqual(payload["laid_in_cost"], 7.5)
        self.assertEqual(payload["landed_cost"], 67.5)


class DailyEmailIngestTests(unittest.TestCase):
    def test_gmail_mailbox_targets_include_all_mail_fallback(self):
        self.assertEqual(
            mailbox_search_targets("imap.gmail.com", "INBOX"),
            ["INBOX", "[Gmail]/All Mail"],
        )

    def test_dedupe_attachments_removes_same_message_attachment(self):
        attachment = AttachmentCandidate("report.xlsx", b"data", None, "message-1", None)

        deduped = dedupe_attachments([attachment, attachment])

        self.assertEqual(deduped, [attachment])

    def test_classify_attachments_uses_filename_keywords(self):
        rb6, rads = classify_attachments(
            [
                AttachmentCandidate("inventory_velocity.xlsx", b"rb6", None, "m1", None),
                AttachmentCandidate("vinosmith_rads.xlsx", b"rads", None, "m2", None),
            ]
        )

        self.assertEqual(rb6.filename, "inventory_velocity.xlsx")
        self.assertEqual(rads.filename, "vinosmith_rads.xlsx")

    def test_storage_path_sanitizes_filename(self):
        self.assertEqual(safe_filename("../Inventory:Velocity.xlsx"), "Inventory_Velocity.xlsx")
        self.assertEqual(
            storage_path(pd.Timestamp("2026-04-30").date(), "rb6_inventory", "../Inventory:Velocity.xlsx"),
            "vinosmith/2026-04-30/rb6_inventory/Inventory_Velocity.xlsx",
        )


if __name__ == "__main__":
    unittest.main()
